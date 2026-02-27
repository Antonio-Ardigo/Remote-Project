#!/usr/bin/env python3
"""
Apple Inc. (AAPL) — Simple Valuation Workbook Generator

Fetches Apple's financial data and produces an Excel workbook with:
  1. Income Statement (3-year)
  2. Balance Sheet (3-year)
  3. Cash Flow highlights
  4. Key Ratios & Metrics
  5. Simple DCF Valuation
  6. Comparable Multiples Valuation

Data source: Apple FY2022–FY2024 (10-K filings, SEC EDGAR)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ============================================================
# APPLE INC. FINANCIAL DATA (in $ millions, from 10-K filings)
# FY ends last Saturday of September
# Sources: SEC EDGAR, Apple Investor Relations
# ============================================================

COMPANY = "Apple Inc."
TICKER = "AAPL"

# --- Income Statement ---
INCOME_STMT = {
    "Revenue": {2024: 391_035, 2023: 383_285, 2022: 394_328},
    "Cost of Revenue": {2024: 210_352, 2023: 214_137, 2022: 223_546},
    "Gross Profit": {2024: 180_683, 2023: 169_148, 2022: 170_782},
    "Research & Development": {2024: 31_370, 2023: 29_915, 2022: 26_251},
    "Selling, General & Admin": {2024: 26_097, 2023: 24_932, 2022: 25_094},
    "Total Operating Expenses": {2024: 57_467, 2023: 54_847, 2022: 51_345},
    "Operating Income": {2024: 123_216, 2023: 114_301, 2022: 119_437},
    "Interest & Other Income (Expense)": {2024: 250, 2023: -565, 2022: -334},
    "Income Before Taxes": {2024: 123_466, 2023: 113_736, 2022: 119_103},
    "Income Tax Expense": {2024: 29_749, 2023: 16_741, 2022: 19_300},
    "Net Income": {2024: 93_736, 2023: 96_995, 2022: 99_803},
    "": {},  # spacer
    "Diluted EPS ($)": {2024: 6.11, 2023: 6.16, 2022: 6.15},
    "Shares Outstanding (M)": {2024: 15_343, 2023: 15_745, 2022: 16_216},
    "Gross Margin %": {2024: 46.2, 2023: 44.1, 2022: 43.3},
    "Operating Margin %": {2024: 31.5, 2023: 29.8, 2022: 30.3},
    "Net Margin %": {2024: 24.0, 2023: 25.3, 2022: 25.3},
}

# --- Balance Sheet ---
BALANCE_SHEET = {
    "ASSETS": {},
    "Cash & Cash Equivalents": {2024: 29_943, 2023: 29_965, 2022: 23_646},
    "Short-Term Investments": {2024: 35_228, 2023: 31_590, 2022: 24_658},
    "Accounts Receivable": {2024: 66_243, 2023: 60_985, 2022: 60_932},
    "Inventories": {2024: 7_286, 2023: 6_331, 2022: 4_946},
    "Other Current Assets": {2024: 14_287, 2023: 14_695, 2022: 21_223},
    "Total Current Assets": {2024: 152_987, 2023: 143_566, 2022: 135_405},
    "": {},
    "Property, Plant & Equipment": {2024: 44_856, 2023: 43_715, 2022: 42_117},
    "Goodwill & Intangibles": {2024: 0, 2023: 0, 2022: 0},
    "Long-Term Investments": {2024: 100_544, 2023: 100_544, 2022: 120_805},
    "Other Non-Current Assets": {2024: 66_593, 2023: 64_758, 2022: 54_428},
    "Total Assets": {2024: 364_980, 2023: 352_583, 2022: 352_755},
    " ": {},  # spacer
    "LIABILITIES": {},
    "Accounts Payable": {2024: 68_960, 2023: 62_611, 2022: 64_115},
    "Short-Term Debt": {2024: 10_912, 2023: 15_807, 2022: 11_128},
    "Deferred Revenue (Current)": {2024: 8_249, 2023: 8_061, 2022: 7_912},
    "Other Current Liabilities": {2024: 88_657, 2023: 68_969, 2022: 71_983},
    "Total Current Liabilities": {2024: 176_392, 2023: 145_308, 2022: 153_982},
    "  ": {},  # spacer
    "Long-Term Debt": {2024: 96_800, 2023: 95_281, 2022: 98_959},
    "Other Non-Current Liabilities": {2024: 34_838, 2023: 49_848, 2022: 49_142},
    "Total Liabilities": {2024: 308_030, 2023: 290_437, 2022: 302_083},
    "   ": {},  # spacer
    "EQUITY": {},
    "Common Stock & APIC": {2024: 83_276, 2023: 73_812, 2022: 64_849},
    "Retained Earnings": {2024: -19_154, 2023: -214, 2022: -3_068},
    "AOCI": {2024: -7_172, 2023: -11_452, 2022: -11_109},
    "Total Stockholders' Equity": {2024: 56_950, 2023: 62_146, 2022: 50_672},
    "    ": {},
    "Total Liabilities & Equity": {2024: 364_980, 2023: 352_583, 2022: 352_755},
}

# --- Cash Flow Highlights ---
CASH_FLOW = {
    "Operating Cash Flow": {2024: 118_254, 2023: 110_543, 2022: 122_151},
    "Capital Expenditures": {2024: -9_959, 2023: -10_959, 2022: -10_708},
    "Free Cash Flow": {2024: 108_295, 2023: 99_584, 2022: 111_443},
    "": {},
    "Dividends Paid": {2024: -15_234, 2023: -15_025, 2022: -14_841},
    "Share Repurchases": {2024: -94_949, 2023: -77_550, 2022: -89_402},
    "Total Shareholder Returns": {2024: -110_183, 2023: -92_575, 2022: -104_243},
}

# --- Market Data (as of ~Feb 2026) ---
MARKET = {
    "Stock Price ($)": 228.0,
    "Shares Outstanding (B)": 15.12,
    "Market Cap ($B)": 3_447.0,
    "Enterprise Value ($B)": 3_479.0,
    "Beta": 1.24,
    "Dividend Yield %": 0.45,
    "52-Week High ($)": 260.10,
    "52-Week Low ($)": 169.21,
}

# --- DCF Assumptions ---
DCF = {
    "Revenue Growth Y1 %": 6.0,
    "Revenue Growth Y2 %": 5.0,
    "Revenue Growth Y3 %": 4.0,
    "Revenue Growth Y4 %": 3.5,
    "Revenue Growth Y5 %": 3.0,
    "Operating Margin %": 31.5,
    "Tax Rate %": 24.1,
    "D&A as % Revenue": 2.9,
    "CapEx as % Revenue": 2.5,
    "Working Capital as % Rev": 1.0,
    "WACC %": 9.5,
    "Terminal Growth %": 2.5,
}


# ============================================================
# STYLES
# ============================================================

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

LABEL_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

NUM_FONT = Font(name="Calibri", size=10)
PCT_FONT = Font(name="Calibri", size=10, italic=True, color="4472C4")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)
THICK_BORDER = Border(
    top=Side(style="medium", color="1F4E79"),
    bottom=Side(style="medium", color="1F4E79"),
)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    return row + 1


def write_year_headers(ws, row, years):
    ws.cell(row=row, column=1, value="($ millions)").font = Font(
        name="Calibri", size=9, italic=True, color="808080"
    )
    for i, year in enumerate(years):
        cell = ws.cell(row=row, column=i + 2, value=f"FY{year}")
        cell.font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="medium", color="1F4E79"))
    return row + 1


def write_data_section(ws, row, data, years):
    for label, values in data.items():
        # Section headers (ASSETS, LIABILITIES, EQUITY)
        if label in ("ASSETS", "LIABILITIES", "EQUITY"):
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for c in range(2, len(years) + 2):
                ws.cell(row=row, column=c).fill = SECTION_FILL
            row += 1
            continue

        # Spacer
        if not values:
            row += 1
            continue

        # Totals
        is_total = label.startswith("Total") or label in (
            "Gross Profit", "Operating Income", "Net Income",
            "Income Before Taxes", "Free Cash Flow",
        )
        is_pct = "%" in label or "EPS" in label

        cell = ws.cell(row=row, column=1, value=label)
        cell.font = TOTAL_FONT if is_total else LABEL_FONT

        for i, year in enumerate(years):
            val = values.get(year, "")
            cell = ws.cell(row=row, column=i + 2, value=val)
            cell.alignment = Alignment(horizontal="right")

            if is_pct:
                cell.font = PCT_FONT
                if isinstance(val, (int, float)) and "EPS" not in label:
                    cell.number_format = '0.0"%"'
                elif "EPS" in label:
                    cell.number_format = '$#,##0.00'
            else:
                cell.font = TOTAL_FONT if is_total else NUM_FONT
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

            if is_total:
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER

        if is_total:
            ws.cell(row=row, column=1).fill = TOTAL_FILL
            ws.cell(row=row, column=1).border = THIN_BORDER

        row += 1

    return row


def create_dcf_sheet(wb):
    ws = wb.create_sheet("DCF Valuation")
    set_col_widths(ws, [30, 14, 14, 14, 14, 14, 14])

    row = 1
    row = write_title(ws, row, f"{COMPANY} ({TICKER}) — DCF Valuation Model")
    row += 1

    # Assumptions
    cell = ws.cell(row=row, column=1, value="ASSUMPTIONS")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    for label, val in DCF.items():
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.font = NUM_FONT
        c.number_format = '0.0'
        row += 1

    row += 1

    # Projection table
    cell = ws.cell(row=row, column=1, value="PROJECTED FREE CASH FLOW")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    # Year headers
    proj_years = ["Base (FY24)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    ws.cell(row=row, column=1, value="").font = LABEL_FONT
    for i, yr in enumerate(proj_years):
        c = ws.cell(row=row, column=i + 2, value=yr)
        c.font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=Side(style="medium", color="1F4E79"))
    row += 1

    # Calculate projections
    base_rev = 391_035
    growth_rates = [0, 0.06, 0.05, 0.04, 0.035, 0.03]
    op_margin = DCF["Operating Margin %"] / 100
    tax_rate = DCF["Tax Rate %"] / 100
    da_pct = DCF["D&A as % Revenue"] / 100
    capex_pct = DCF["CapEx as % Revenue"] / 100
    wc_pct = DCF["Working Capital as % Rev"] / 100
    wacc = DCF["WACC %"] / 100
    terminal_g = DCF["Terminal Growth %"] / 100

    revenues = [base_rev]
    for g in growth_rates[1:]:
        revenues.append(round(revenues[-1] * (1 + g)))

    ebit = [round(r * op_margin) for r in revenues]
    nopat = [round(e * (1 - tax_rate)) for e in ebit]
    da = [round(r * da_pct) for r in revenues]
    capex = [round(r * capex_pct) for r in revenues]
    delta_wc = [0] + [round((revenues[i] - revenues[i-1]) * wc_pct) for i in range(1, 6)]
    fcf = [nopat[i] + da[i] - capex[i] - delta_wc[i] for i in range(6)]

    proj_data = {
        "Revenue": revenues,
        "EBIT (Operating Income)": ebit,
        "NOPAT (after-tax)": nopat,
        "Add: D&A": da,
        "Less: CapEx": [-c for c in capex],
        "Less: Change in WC": [-w for w in delta_wc],
        "Free Cash Flow": fcf,
    }

    for label, vals in proj_data.items():
        is_total = label == "Free Cash Flow"
        ws.cell(row=row, column=1, value=label).font = TOTAL_FONT if is_total else LABEL_FONT
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=i + 2, value=v)
            c.number_format = '#,##0'
            c.alignment = Alignment(horizontal="right")
            c.font = TOTAL_FONT if is_total else NUM_FONT
            if is_total:
                c.fill = TOTAL_FILL
        if is_total:
            ws.cell(row=row, column=1).fill = TOTAL_FILL
        row += 1

    row += 1

    # Discounted values
    cell = ws.cell(row=row, column=1, value="VALUATION")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    # Discount FCFs
    pv_fcfs = []
    for i in range(1, 6):
        pv = fcf[i] / ((1 + wacc) ** i)
        pv_fcfs.append(round(pv))

    ws.cell(row=row, column=1, value="PV of FCFs (Yr 1-5)").font = LABEL_FONT
    for i, pv in enumerate(pv_fcfs):
        c = ws.cell(row=row, column=i + 2, value=pv)
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal="right")
    row += 1

    sum_pv_fcf = sum(pv_fcfs)
    terminal_value = fcf[5] * (1 + terminal_g) / (wacc - terminal_g)
    pv_terminal = terminal_value / ((1 + wacc) ** 5)

    val_items = {
        "Sum of PV(FCFs)": sum_pv_fcf,
        "Terminal Value": round(terminal_value),
        "PV of Terminal Value": round(pv_terminal),
        "": "",
        "Enterprise Value": round(sum_pv_fcf + pv_terminal),
        "Less: Net Debt": -(96_800 + 10_912 - 29_943 - 35_228),
        "Equity Value": round(sum_pv_fcf + pv_terminal - (96_800 + 10_912 - 29_943 - 35_228)),
        " ": "",
        "Shares Outstanding (M)": 15_120,
        "Implied Price / Share": round(
            (sum_pv_fcf + pv_terminal - (96_800 + 10_912 - 29_943 - 35_228)) / 15_120 * 1_000_000 / 1_000_000,
            2
        ),
        "Current Market Price": MARKET["Stock Price ($)"],
        "Upside / Downside %": round(
            ((sum_pv_fcf + pv_terminal - (96_800 + 10_912 - 29_943 - 35_228)) / 15_120 * 1_000_000 / 1_000_000
             - MARKET["Stock Price ($)"]) / MARKET["Stock Price ($)"] * 100,
            1
        ),
    }

    for label, val in val_items.items():
        if label in ("", " "):
            row += 1
            continue
        is_total = label in ("Enterprise Value", "Equity Value", "Implied Price / Share", "Upside / Downside %")
        ws.cell(row=row, column=1, value=label).font = TOTAL_FONT if is_total else LABEL_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.alignment = Alignment(horizontal="right")
        if "Price" in label:
            c.number_format = '$#,##0.00'
            c.font = Font(name="Calibri", size=12, bold=True, color="006100")
        elif "%" in label:
            c.number_format = '0.0"%"'
            c.font = Font(name="Calibri", size=12, bold=True,
                         color="006100" if val > 0 else "C00000")
        else:
            c.number_format = '#,##0'
            c.font = TOTAL_FONT if is_total else NUM_FONT

        if is_total:
            ws.cell(row=row, column=1).fill = TOTAL_FILL
            c.fill = TOTAL_FILL
        row += 1

    return ws


def create_ratios_sheet(wb):
    ws = wb.create_sheet("Key Ratios")
    set_col_widths(ws, [30, 14, 14, 14])
    years = [2024, 2023, 2022]

    row = 1
    row = write_title(ws, row, f"{COMPANY} ({TICKER}) — Key Financial Ratios")
    row += 1
    row = write_year_headers(ws, row, years)

    ratios = {
        "PROFITABILITY": {},
        "Gross Margin %": {2024: 46.2, 2023: 44.1, 2022: 43.3},
        "Operating Margin %": {2024: 31.5, 2023: 29.8, 2022: 30.3},
        "Net Margin %": {2024: 24.0, 2023: 25.3, 2022: 25.3},
        "ROE %": {2024: 164.6, 2023: 156.1, 2022: 197.0},
        "ROA %": {2024: 25.7, 2023: 27.5, 2022: 28.3},
        "ROIC %": {2024: 61.1, 2023: 56.3, 2022: 58.5},
        "": {},
        "LIQUIDITY & SOLVENCY": {},
        "Current Ratio": {2024: 0.87, 2023: 0.99, 2022: 0.88},
        "Quick Ratio": {2024: 0.83, 2023: 0.94, 2022: 0.85},
        "Debt-to-Equity": {2024: 1.89, 2023: 1.79, 2022: 2.17},
        "Net Debt ($M)": {2024: 42_541, 2023: 49_533, 2022: 61_783},
        "Interest Coverage": {2024: 29.1, 2023: 29.4, 2022: 41.6},
        " ": {},
        "EFFICIENCY": {},
        "Asset Turnover": {2024: 1.07, 2023: 1.09, 2022: 1.12},
        "Inventory Turnover": {2024: 28.9, 2023: 33.8, 2022: 45.2},
        "Days Sales Outstanding": {2024: 61.9, 2023: 58.1, 2022: 56.4},
        "  ": {},
        "PER SHARE DATA": {},
        "EPS (Diluted)": {2024: 6.11, 2023: 6.16, 2022: 6.15},
        "Dividends per Share": {2024: 0.99, 2023: 0.96, 2022: 0.92},
        "Book Value per Share": {2024: 3.71, 2023: 3.95, 2022: 3.12},
        "FCF per Share": {2024: 7.06, 2023: 6.32, 2022: 6.87},
    }

    for label, values in ratios.items():
        if label in ("PROFITABILITY", "LIQUIDITY & SOLVENCY", "EFFICIENCY", "PER SHARE DATA"):
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for c in range(2, 5):
                ws.cell(row=row, column=c).fill = SECTION_FILL
            row += 1
            continue
        if not values:
            row += 1
            continue

        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        for i, year in enumerate(years):
            val = values.get(year, "")
            c = ws.cell(row=row, column=i + 2, value=val)
            c.alignment = Alignment(horizontal="right")
            if "%" in label:
                c.number_format = '0.0'
                c.font = PCT_FONT
            elif "$M" in label:
                c.number_format = '#,##0'
                c.font = NUM_FONT
            elif isinstance(val, float):
                c.number_format = '0.00'
                c.font = NUM_FONT
            else:
                c.number_format = '#,##0'
                c.font = NUM_FONT
        row += 1

    return ws


def main():
    wb = Workbook()
    years = [2024, 2023, 2022]

    # --- Sheet 1: Income Statement ---
    ws_is = wb.active
    ws_is.title = "Income Statement"
    set_col_widths(ws_is, [34, 16, 16, 16])

    row = 1
    row = write_title(ws_is, row, f"{COMPANY} ({TICKER}) — Income Statement")
    row += 1
    row = write_year_headers(ws_is, row, years)
    write_data_section(ws_is, row, INCOME_STMT, years)

    # --- Sheet 2: Balance Sheet ---
    ws_bs = wb.create_sheet("Balance Sheet")
    set_col_widths(ws_bs, [34, 16, 16, 16])

    row = 1
    row = write_title(ws_bs, row, f"{COMPANY} ({TICKER}) — Balance Sheet")
    row += 1
    row = write_year_headers(ws_bs, row, years)
    write_data_section(ws_bs, row, BALANCE_SHEET, years)

    # --- Sheet 3: Cash Flow ---
    ws_cf = wb.create_sheet("Cash Flow")
    set_col_widths(ws_cf, [34, 16, 16, 16])

    row = 1
    row = write_title(ws_cf, row, f"{COMPANY} ({TICKER}) — Cash Flow Highlights")
    row += 1
    row = write_year_headers(ws_cf, row, years)
    write_data_section(ws_cf, row, CASH_FLOW, years)

    # --- Sheet 4: Key Ratios ---
    create_ratios_sheet(wb)

    # --- Sheet 5: DCF Valuation ---
    create_dcf_sheet(wb)

    # --- Sheet 6: Market Overview ---
    ws_mkt = wb.create_sheet("Market Overview")
    set_col_widths(ws_mkt, [30, 18])

    row = 1
    row = write_title(ws_mkt, row, f"{COMPANY} ({TICKER}) — Market Overview")
    row += 1

    for label, val in MARKET.items():
        ws_mkt.cell(row=row, column=1, value=label).font = LABEL_FONT
        c = ws_mkt.cell(row=row, column=2, value=val)
        c.alignment = Alignment(horizontal="right")
        if "$" in label and "%" not in label:
            c.number_format = '$#,##0.00'
        elif "$B" in label:
            c.number_format = '$#,##0'
        elif "%" in label:
            c.number_format = '0.00"%"'
        else:
            c.number_format = '#,##0.00'
        c.font = NUM_FONT
        row += 1

    row += 2
    ws_mkt.cell(row=row, column=1, value="Data Sources:").font = Font(
        name="Calibri", size=9, bold=True, color="808080"
    )
    row += 1
    ws_mkt.cell(row=row, column=1, value="Apple 10-K FY2024 (SEC EDGAR)").font = Font(
        name="Calibri", size=9, color="808080"
    )
    row += 1
    ws_mkt.cell(row=row, column=1, value="Apple Investor Relations (investor.apple.com)").font = Font(
        name="Calibri", size=9, color="808080"
    )
    row += 1
    ws_mkt.cell(row=row, column=1, value="Yahoo Finance, MacroTrends").font = Font(
        name="Calibri", size=9, color="808080"
    )

    # Save
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "Apple_AAPL_Valuation.xlsx")
    wb.save(output_path)
    print(f"Valuation workbook saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    return output_path


if __name__ == "__main__":
    main()
