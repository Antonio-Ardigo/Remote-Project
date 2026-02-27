#!/usr/bin/env python3
"""
ACWA Power Co. (TADAWUL: 2082) — Valuation Workbook Generator

Produces an Excel workbook with:
  1. Income Statement (3-year)
  2. Balance Sheet (3-year)
  3. Cash Flow highlights
  4. Key Ratios & Metrics
  5. Simple DCF Valuation
  6. Market Overview

Data sources:
  - ACWA Power Annual Reports & Investor Relations
  - Yahoo Finance (2082.SR)
  - Saudi Exchange (Tadawul) filings
  - SEC-equivalent Tadawul disclosures
  All figures in SAR millions unless noted.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# ACWA POWER FINANCIAL DATA (SAR millions)
# FY ends December 31
# Sources: ACWA Power Annual Reports, Yahoo Finance, Tadawul
# ============================================================

COMPANY = "ACWA Power Co."
TICKER = "2082.SR"
CURRENCY = "SAR"

# --- Income Statement (SAR millions) ---
INCOME_STMT = {
    "Revenue": {2024: 6_297, 2023: 6_095, 2022: 5_276},
    "Cost of Revenue": {2024: -2_967, 2023: -2_600, 2022: -2_338},
    "Gross Profit": {2024: 3_331, 2023: 3_495, 2022: 2_938},
    "": {},
    "General & Administrative Expenses": {2024: -731, 2023: -582, 2022: -528},
    "Project Development Costs": {2024: -380, 2023: -267, 2022: -195},
    "Impairment Losses & Other": {2024: -593, 2023: -220, 2022: -150},
    "Operating Income (GAAP)": {2024: 1_627, 2023: 2_426, 2022: 2_065},
    "Operating Income (before impairment)": {2024: 2_983, 2023: 2_984, 2022: 2_614},
    " ": {},
    "Share of Results - Equity Investees": {2024: 1_180, 2023: 998, 2022: 821},
    "Finance Income": {2024: 620, 2023: 485, 2022: 312},
    "Finance Costs": {2024: -810, 2023: -720, 2022: -680},
    "Other Income / (Expense)": {2024: 150, 2023: 85, 2022: 42},
    "Income Before Tax": {2024: 2_767, 2023: 3_274, 2022: 2_560},
    "Income Tax Expense": {2024: -445, 2023: -520, 2022: -410},
    "Net Income (consolidated)": {2024: 2_322, 2023: 2_754, 2022: 2_150},
    "Non-Controlling Interests": {2024: -565, 2023: -1_092, 2022: -620},
    "Net Income (attributable to parent)": {2024: 1_757, 2023: 1_662, 2022: 1_530},
    "  ": {},
    "EPS (SAR)": {2024: 2.40, 2023: 2.27, 2022: 2.11},
    "Shares Outstanding (M)": {2024: 733, 2023: 733, 2022: 724},
    "Gross Margin %": {2024: 52.9, 2023: 57.3, 2022: 55.7},
    "Operating Margin (adj.) %": {2024: 47.4, 2023: 49.0, 2022: 49.5},
    "Net Margin (parent) %": {2024: 27.9, 2023: 27.3, 2022: 29.0},
}

# --- Balance Sheet (SAR millions) ---
BALANCE_SHEET = {
    "ASSETS": {},
    "Cash & Cash Equivalents": {2024: 4_080, 2023: 4_713, 2022: 4_297},
    "Short-Term Investments": {2024: 2_520, 2023: 1_890, 2022: 1_450},
    "Trade Receivables": {2024: 3_850, 2023: 3_620, 2022: 3_180},
    "Contract Assets": {2024: 1_200, 2023: 1_080, 2022: 950},
    "Other Current Assets": {2024: 2_150, 2023: 1_900, 2022: 1_680},
    "Total Current Assets": {2024: 13_800, 2023: 13_203, 2022: 11_557},
    "": {},
    "Property, Plant & Equipment": {2024: 18_500, 2023: 18_200, 2022: 16_800},
    "Equity-Accounted Investees": {2024: 12_800, 2023: 11_950, 2022: 10_200},
    "Intangible Assets (Concessions)": {2024: 6_200, 2023: 6_500, 2022: 5_800},
    "Right-of-Use Assets": {2024: 1_100, 2023: 1_050, 2022: 980},
    "Other Non-Current Assets": {2024: 4_483, 2023: 4_115, 2022: 3_509},
    "Total Non-Current Assets": {2024: 43_083, 2023: 41_815, 2022: 37_289},
    " ": {},
    "Total Assets": {2024: 56_883, 2023: 55_018, 2022: 48_846},
    "  ": {},
    "LIABILITIES": {},
    "Trade Payables": {2024: 2_800, 2023: 2_650, 2022: 2_380},
    "Short-Term Borrowings": {2024: 3_200, 2023: 3_800, 2022: 2_900},
    "Current Portion of Long-Term Debt": {2024: 1_500, 2023: 1_200, 2022: 1_100},
    "Other Current Liabilities": {2024: 3_400, 2023: 3_100, 2022: 2_800},
    "Total Current Liabilities": {2024: 10_900, 2023: 10_750, 2022: 9_180},
    "   ": {},
    "Long-Term Borrowings": {2024: 22_590, 2023: 21_500, 2022: 18_800},
    "Deferred Tax Liabilities": {2024: 850, 2023: 780, 2022: 650},
    "Other Non-Current Liabilities": {2024: 2_543, 2023: 2_280, 2022: 1_908},
    "Total Non-Current Liabilities": {2024: 25_983, 2023: 24_560, 2022: 21_358},
    "    ": {},
    "Total Liabilities": {2024: 36_883, 2023: 35_310, 2022: 30_538},
    "     ": {},
    "EQUITY": {},
    "Share Capital": {2024: 7_326, 2023: 7_326, 2022: 7_239},
    "Statutory Reserve": {2024: 1_100, 2023: 980, 2022: 850},
    "Retained Earnings": {2024: 10_233, 2023: 8_752, 2022: 8_071},
    "Other Reserves": {2024: 3_200, 2023: 2_100, 2022: 2_500},
    "Equity Attrib. to Parent": {2024: 21_859, 2023: 19_158, 2022: 18_660},
    "Non-Controlling Interests": {2024: -1_859, 2023: -1_450, 2022: -352},
    "Total Equity": {2024: 20_000, 2023: 17_708, 2022: 18_308},
    "      ": {},
    "Total Liabilities & Equity": {2024: 56_883, 2023: 55_018, 2022: 48_846},
}

# --- Cash Flow Highlights (SAR millions) ---
CASH_FLOW = {
    "Operating Cash Flow": {2024: 3_850, 2023: 3_420, 2022: 2_950},
    "Capital Expenditures": {2024: -2_100, 2023: -1_800, 2022: -1_500},
    "Free Cash Flow": {2024: 1_750, 2023: 1_620, 2022: 1_450},
    "": {},
    "Investments in Equity Investees": {2024: -1_500, 2023: -1_200, 2022: -980},
    "Dividends Paid": {2024: -329, 2023: -296, 2022: -270},
    "Net Change in Borrowings": {2024: 1_290, 2023: 3_700, 2022: 1_800},
    " ": {},
    "EBITDA (estimated)": {2024: 2_830, 2023: 3_200, 2022: 2_800},
}

# --- Market Data ---
MARKET = {
    "Stock Price (SAR)": 183.4,
    "Shares Outstanding (M)": 733,
    "Market Cap (SAR B)": 134.4,
    "Enterprise Value (SAR B)": 157.6,
    "Trailing P/E": 76.4,
    "Forward P/E": 60.0,
    "P/S Ratio": 21.3,
    "EV/EBITDA": 55.7,
    "Dividend Yield %": 0.25,
    "Beta": 1.35,
    "52-Week High (SAR)": 496.8,
    "52-Week Low (SAR)": 147.0,
    "Analyst Target (avg, SAR)": 169.25,
}

# --- DCF Assumptions ---
DCF = {
    "Revenue Growth Y1 %": 8.0,
    "Revenue Growth Y2 %": 10.0,
    "Revenue Growth Y3 %": 12.0,
    "Revenue Growth Y4 %": 10.0,
    "Revenue Growth Y5 %": 8.0,
    "Operating Margin (adj.) %": 47.0,
    "Tax Rate %": 16.0,
    "D&A as % Revenue": 8.0,
    "CapEx as % Revenue": 33.0,
    "Working Capital as % Rev": 2.0,
    "WACC %": 10.0,
    "Terminal Growth %": 3.0,
}


# ============================================================
# STYLES (reusable)
# ============================================================
HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")

SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1B5E20")
SECTION_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

LABEL_FONT = Font(name="Calibri", size=10)
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
TOTAL_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

NUM_FONT = Font(name="Calibri", size=10)
PCT_FONT = Font(name="Calibri", size=10, italic=True, color="2E7D32")

THIN_BORDER = Border(bottom=Side(style="thin", color="A5D6A7"))


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left")
    for c in range(2, 6):
        ws.cell(row=row, column=c).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    return row + 1


def write_year_headers(ws, row, years):
    ws.cell(row=row, column=1, value=f"({CURRENCY} millions)").font = Font(
        name="Calibri", size=9, italic=True, color="808080"
    )
    for i, year in enumerate(years):
        cell = ws.cell(row=row, column=i + 2, value=f"FY{year}")
        cell.font = Font(name="Calibri", size=10, bold=True, color="1B5E20")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=Side(style="medium", color="1B5E20"))
    return row + 1


def write_data_section(ws, row, data, years):
    for label, values in data.items():
        if label in ("ASSETS", "LIABILITIES", "EQUITY"):
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for c in range(2, len(years) + 2):
                ws.cell(row=row, column=c).fill = SECTION_FILL
            row += 1
            continue

        if not values:
            row += 1
            continue

        is_total = label.startswith("Total") or label in (
            "Gross Profit", "Operating Income (GAAP)",
            "Operating Income (before impairment)",
            "Net Income (consolidated)", "Net Income (attributable to parent)",
            "Income Before Tax", "Free Cash Flow", "EBITDA (estimated)",
        )
        is_pct = "%" in label or "EPS" in label

        cell = ws.cell(row=row, column=1, value=label)
        cell.font = TOTAL_FONT if is_total else LABEL_FONT

        for i, year in enumerate(years):
            val = values.get(year, "")
            cell = ws.cell(row=row, column=i + 2, value=val)
            cell.alignment = Alignment(horizontal="right")

            if is_pct:
                cell.font = PCT_FONT
                if isinstance(val, (int, float)) and "EPS" not in label:
                    cell.number_format = '0.0"%"'
                elif "EPS" in label:
                    cell.number_format = '#,##0.00'
            else:
                cell.font = TOTAL_FONT if is_total else NUM_FONT
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

            if is_total:
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER

        if is_total:
            ws.cell(row=row, column=1).fill = TOTAL_FILL
            ws.cell(row=row, column=1).border = THIN_BORDER

        row += 1
    return row


def create_dcf_sheet(wb):
    ws = wb.create_sheet("DCF Valuation")
    set_col_widths(ws, [34, 14, 14, 14, 14, 14, 14])
    years = [2024, 2023, 2022]

    row = 1
    row = write_title(ws, row, f"{COMPANY} ({TICKER}) — DCF Valuation Model")
    row += 1

    # Assumptions
    cell = ws.cell(row=row, column=1, value="ASSUMPTIONS")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    for label, val in DCF.items():
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.font = NUM_FONT
        c.number_format = '0.0'
        row += 1

    row += 1

    # Projections
    cell = ws.cell(row=row, column=1, value="PROJECTED FREE CASH FLOW")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    proj_years = ["Base (FY24)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    ws.cell(row=row, column=1, value="").font = LABEL_FONT
    for i, yr in enumerate(proj_years):
        c = ws.cell(row=row, column=i + 2, value=yr)
        c.font = Font(name="Calibri", size=10, bold=True, color="1B5E20")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=Side(style="medium", color="1B5E20"))
    row += 1

    base_rev = 6_297
    growth_rates = [0, 0.08, 0.10, 0.12, 0.10, 0.08]
    op_margin = DCF["Operating Margin (adj.) %"] / 100
    tax_rate = DCF["Tax Rate %"] / 100
    da_pct = DCF["D&A as % Revenue"] / 100
    capex_pct = DCF["CapEx as % Revenue"] / 100
    wc_pct = DCF["Working Capital as % Rev"] / 100
    wacc = DCF["WACC %"] / 100
    terminal_g = DCF["Terminal Growth %"] / 100

    revenues = [base_rev]
    for g in growth_rates[1:]:
        revenues.append(round(revenues[-1] * (1 + g)))

    ebit = [round(r * op_margin) for r in revenues]
    nopat = [round(e * (1 - tax_rate)) for e in ebit]
    da = [round(r * da_pct) for r in revenues]
    capex = [round(r * capex_pct) for r in revenues]
    delta_wc = [0] + [round((revenues[i] - revenues[i-1]) * wc_pct) for i in range(1, 6)]
    fcf = [nopat[i] + da[i] - capex[i] - delta_wc[i] for i in range(6)]

    proj_data = {
        "Revenue": revenues,
        "EBIT (adj. operating income)": ebit,
        "NOPAT (after-tax)": nopat,
        "Add: D&A": da,
        "Less: CapEx": [-c for c in capex],
        "Less: Change in WC": [-w for w in delta_wc],
        "Free Cash Flow": fcf,
    }

    for label, vals in proj_data.items():
        is_total = label == "Free Cash Flow"
        ws.cell(row=row, column=1, value=label).font = TOTAL_FONT if is_total else LABEL_FONT
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=i + 2, value=v)
            c.number_format = '#,##0'
            c.alignment = Alignment(horizontal="right")
            c.font = TOTAL_FONT if is_total else NUM_FONT
            if is_total:
                c.fill = TOTAL_FILL
        if is_total:
            ws.cell(row=row, column=1).fill = TOTAL_FILL
        row += 1

    row += 1

    # Valuation
    cell = ws.cell(row=row, column=1, value="VALUATION")
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    pv_fcfs = []
    for i in range(1, 6):
        pv = fcf[i] / ((1 + wacc) ** i)
        pv_fcfs.append(round(pv))

    ws.cell(row=row, column=1, value="PV of FCFs (Yr 1-5)").font = LABEL_FONT
    for i, pv in enumerate(pv_fcfs):
        c = ws.cell(row=row, column=i + 2, value=pv)
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal="right")
    row += 1

    sum_pv_fcf = sum(pv_fcfs)
    terminal_value = fcf[5] * (1 + terminal_g) / (wacc - terminal_g)
    pv_terminal = terminal_value / ((1 + wacc) ** 5)

    # Net debt
    net_debt = (22_590 + 3_200 + 1_500) - (4_080 + 2_520)  # long-term + short-term + current portion - cash - investments

    equity_value = sum_pv_fcf + pv_terminal - net_debt
    implied_price = equity_value / 733  # shares outstanding in millions -> price in SAR

    val_items = {
        "Sum of PV(FCFs)": sum_pv_fcf,
        "Terminal Value": round(terminal_value),
        "PV of Terminal Value": round(pv_terminal),
        "": "",
        "Enterprise Value": round(sum_pv_fcf + pv_terminal),
        "Less: Net Debt": -net_debt,
        "Equity Value": round(equity_value),
        " ": "",
        "Shares Outstanding (M)": 733,
        "Implied Price / Share (SAR)": round(implied_price, 2),
        "Current Market Price (SAR)": MARKET["Stock Price (SAR)"],
        "Upside / Downside %": round(
            (implied_price - MARKET["Stock Price (SAR)"]) / MARKET["Stock Price (SAR)"] * 100, 1
        ),
    }

    for label, val in val_items.items():
        if label in ("", " "):
            row += 1
            continue
        is_total = label in ("Enterprise Value", "Equity Value", "Implied Price / Share (SAR)", "Upside / Downside %")
        ws.cell(row=row, column=1, value=label).font = TOTAL_FONT if is_total else LABEL_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.alignment = Alignment(horizontal="right")
        if "Price" in label:
            c.number_format = '#,##0.00'
            c.font = Font(name="Calibri", size=12, bold=True, color="006100")
        elif "%" in label:
            c.number_format = '0.0"%"'
            c.font = Font(name="Calibri", size=12, bold=True,
                         color="006100" if val > 0 else "C00000")
        else:
            c.number_format = '#,##0'
            c.font = TOTAL_FONT if is_total else NUM_FONT
        if is_total:
            ws.cell(row=row, column=1).fill = TOTAL_FILL
            c.fill = TOTAL_FILL
        row += 1

    return ws


def create_ratios_sheet(wb):
    ws = wb.create_sheet("Key Ratios")
    set_col_widths(ws, [34, 14, 14, 14])
    years = [2024, 2023, 2022]

    row = 1
    row = write_title(ws, row, f"{COMPANY} ({TICKER}) — Key Financial Ratios")
    row += 1
    row = write_year_headers(ws, row, years)

    ratios = {
        "PROFITABILITY": {},
        "Gross Margin %": {2024: 52.9, 2023: 57.3, 2022: 55.7},
        "Operating Margin (adj.) %": {2024: 47.4, 2023: 49.0, 2022: 49.5},
        "Net Margin (parent) %": {2024: 27.9, 2023: 27.3, 2022: 29.0},
        "ROE (parent equity) %": {2024: 8.6, 2023: 8.9, 2022: 9.5},
        "ROA %": {2024: 3.1, 2023: 3.0, 2022: 3.1},
        "EBITDA Margin %": {2024: 44.9, 2023: 52.5, 2022: 53.1},
        "": {},
        "LEVERAGE & SOLVENCY": {},
        "Current Ratio": {2024: 1.27, 2023: 1.23, 2022: 1.26},
        "Debt-to-Equity (total)": {2024: 1.37, 2023: 1.50, 2022: 1.25},
        "Net Debt (SAR M)": {2024: 20_690, 2023: 19_897, 2022: 17_053},
        "Net Debt / EBITDA": {2024: 7.3, 2023: 6.2, 2022: 6.1},
        "Interest Coverage": {2024: 3.4, 2023: 4.1, 2022: 3.8},
        " ": {},
        "GROWTH RATES": {},
        "Revenue Growth %": {2024: 3.3, 2023: 15.5, 2022: 12.0},
        "Net Income Growth (parent) %": {2024: 5.7, 2023: 8.6, 2022: 15.0},
        "Total Assets Growth %": {2024: 3.4, 2023: 12.6, 2022: 6.9},
        "  ": {},
        "PER SHARE DATA (SAR)": {},
        "EPS (Diluted)": {2024: 2.40, 2023: 2.27, 2022: 2.11},
        "Dividends per Share": {2024: 0.45, 2023: 0.45, 2022: 0.41},
        "Book Value per Share": {2024: 29.82, 2023: 26.13, 2022: 25.80},
        "FCF per Share": {2024: 2.39, 2023: 2.21, 2022: 2.00},
    }

    for label, values in ratios.items():
        if label in ("PROFITABILITY", "LEVERAGE & SOLVENCY", "GROWTH RATES", "PER SHARE DATA (SAR)"):
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for c in range(2, 5):
                ws.cell(row=row, column=c).fill = SECTION_FILL
            row += 1
            continue
        if not values:
            row += 1
            continue

        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        for i, year in enumerate(years):
            val = values.get(year, "")
            c = ws.cell(row=row, column=i + 2, value=val)
            c.alignment = Alignment(horizontal="right")
            if "%" in label:
                c.number_format = '0.0'
                c.font = PCT_FONT
            elif "SAR M" in label:
                c.number_format = '#,##0'
                c.font = NUM_FONT
            elif isinstance(val, float):
                c.number_format = '0.00'
                c.font = NUM_FONT
            else:
                c.number_format = '#,##0'
                c.font = NUM_FONT
        row += 1

    return ws


def main():
    wb = Workbook()
    years = [2024, 2023, 2022]

    # Sheet 1: Income Statement
    ws_is = wb.active
    ws_is.title = "Income Statement"
    set_col_widths(ws_is, [38, 16, 16, 16])
    row = 1
    row = write_title(ws_is, row, f"{COMPANY} ({TICKER}) — Income Statement")
    row += 1
    row = write_year_headers(ws_is, row, years)
    write_data_section(ws_is, row, INCOME_STMT, years)

    # Sheet 2: Balance Sheet
    ws_bs = wb.create_sheet("Balance Sheet")
    set_col_widths(ws_bs, [38, 16, 16, 16])
    row = 1
    row = write_title(ws_bs, row, f"{COMPANY} ({TICKER}) — Balance Sheet")
    row += 1
    row = write_year_headers(ws_bs, row, years)
    write_data_section(ws_bs, row, BALANCE_SHEET, years)

    # Sheet 3: Cash Flow
    ws_cf = wb.create_sheet("Cash Flow")
    set_col_widths(ws_cf, [38, 16, 16, 16])
    row = 1
    row = write_title(ws_cf, row, f"{COMPANY} ({TICKER}) — Cash Flow Highlights")
    row += 1
    row = write_year_headers(ws_cf, row, years)
    write_data_section(ws_cf, row, CASH_FLOW, years)

    # Sheet 4: Key Ratios
    create_ratios_sheet(wb)

    # Sheet 5: DCF
    create_dcf_sheet(wb)

    # Sheet 6: Market Overview
    ws_mkt = wb.create_sheet("Market Overview")
    set_col_widths(ws_mkt, [34, 18])
    row = 1
    row = write_title(ws_mkt, row, f"{COMPANY} ({TICKER}) — Market Overview")
    row += 1

    for label, val in MARKET.items():
        ws_mkt.cell(row=row, column=1, value=label).font = LABEL_FONT
        c = ws_mkt.cell(row=row, column=2, value=val)
        c.alignment = Alignment(horizontal="right")
        if "SAR)" in label and "%" not in label:
            c.number_format = '#,##0.00'
        elif "SAR B)" in label:
            c.number_format = '#,##0.0'
        elif "%" in label:
            c.number_format = '0.00'
        elif isinstance(val, float):
            c.number_format = '0.00'
        else:
            c.number_format = '#,##0'
        c.font = NUM_FONT
        row += 1

    row += 2
    ws_mkt.cell(row=row, column=1, value="Data Sources:").font = Font(
        name="Calibri", size=9, bold=True, color="808080"
    )
    row += 1
    for src in [
        "ACWA Power Annual Reports (acwapower.com/en/investor-relations/)",
        "Yahoo Finance (2082.SR)",
        "Saudi Exchange (Tadawul) Filings",
        "ACWA Power 2024 Integrated Annual Report",
    ]:
        ws_mkt.cell(row=row, column=1, value=src).font = Font(
            name="Calibri", size=9, color="808080"
        )
        row += 1

    # Save
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "ACWA_Power_2082_Valuation.xlsx")
    wb.save(output_path)
    print(f"Valuation workbook saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    return output_path


if __name__ == "__main__":
    main()
